"""
NSE NFO Lot Size & Price Fetcher
Data source: Yahoo Finance (yfinance) — works 100% from GitHub Actions
Lot sizes: hardcoded from NSE circular (updated periodically)
"""

import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, json

OUT = os.path.join(os.path.dirname(__file__), "..", "data", "nfo_lot_sizes.xlsx")

NAVY  = "1F3864"; GOLD  = "C9A84C"; LIGHT = "EEF2FF"
WHITE = "FFFFFF"; GREEN = "E2EFDA"; ORANGE = "FCE4D6"

# ── NSE F&O Lot Sizes (as per NSE circulars, Apr 2026) ──────────────────────
# Format: (NSE_SYMBOL, Yahoo_SYMBOL, Lot_Size)
FO_STOCKS = [
    ("AARTIIND",    "AARTIIND.NS",    1200),
    ("ABB",         "ABB.NS",          250),
    ("ABBOTINDIA",  "ABBOTINDIA.NS",    40),
    ("ABCAPITAL",   "ABCAPITAL.NS",   3200),
    ("ABFRL",       "ABFRL.NS",       2800),
    ("ACC",         "ACC.NS",          500),
    ("ADANIENT",    "ADANIENT.NS",     875),
    ("ADANIPORTS",  "ADANIPORTS.NS",  1250),
    ("ALKEM",       "ALKEM.NS",        200),
    ("AMBUJACEM",   "AMBUJACEM.NS",   2000),
    ("ANGELONE",    "ANGELONE.NS",     300),
    ("APLAPOLLO",   "APLAPOLLO.NS",    750),
    ("APOLLOHOSP",  "APOLLOHOSP.NS",   125),
    ("APOLLOTYRE",  "APOLLOTYRE.NS",  2700),
    ("ASHOKLEY",    "ASHOKLEY.NS",    5500),
    ("ASIANPAINT",  "ASIANPAINT.NS",   200),
    ("ASTRAL",      "ASTRAL.NS",       400),
    ("ATGL",        "ATGL.NS",         750),
    ("ATUL",        "ATUL.NS",         125),
    ("AUBANK",      "AUBANK.NS",      1000),
    ("AUROPHARMA",  "AUROPHARMA.NS",   650),
    ("AXISBANK",    "AXISBANK.NS",     625),
    ("BAJAJ-AUTO",  "BAJAJ-AUTO.NS",   250),
    ("BAJAJFINSV",  "BAJAJFINSV.NS",   500),
    ("BAJFINANCE",  "BAJFINANCE.NS",   125),
    ("BALKRISIND",  "BALKRISIND.NS",   400),
    ("BALRAMCHIN",  "BALRAMCHIN.NS",  1600),
    ("BANDHANBNK",  "BANDHANBNK.NS",  3600),
    ("BANKBARODA",  "BANKBARODA.NS",  5850),
    ("BEL",         "BEL.NS",         3750),
    ("BERGEPAINT",  "BERGEPAINT.NS",  1100),
    ("BHARTIARTL",  "BHARTIARTL.NS",   500),
    ("BHEL",        "BHEL.NS",        10500),
    ("BIOCON",      "BIOCON.NS",      2900),
    ("BOSCHLTD",    "BOSCHLTD.NS",      50),
    ("BPCL",        "BPCL.NS",        1800),
    ("BRITANNIA",   "BRITANNIA.NS",    200),
    ("BSOFT",       "BSOFT.NS",       2200),
    ("CANBK",       "CANBK.NS",       4350),
    ("CANFINHOME",  "CANFINHOME.NS",   975),
    ("CHAMBLFERT",  "CHAMBLFERT.NS",  2400),
    ("CHOLAFIN",    "CHOLAFIN.NS",     700),
    ("CIPLA",       "CIPLA.NS",        650),
    ("COALINDIA",   "COALINDIA.NS",   2100),
    ("COFORGE",     "COFORGE.NS",      200),
    ("COLPAL",      "COLPAL.NS",       350),
    ("CONCOR",      "CONCOR.NS",      1250),
    ("COROMANDEL",  "COROMANDEL.NS",   500),
    ("CROMPTON",    "CROMPTON.NS",    3000),
    ("CUB",         "CUB.NS",         4700),
    ("CUMMINSIND",  "CUMMINSIND.NS",   600),
    ("DABUR",       "DABUR.NS",       2750),
    ("DALBHARAT",   "DALBHARAT.NS",    300),
    ("DEEPAKNTR",   "DEEPAKNTR.NS",    300),
    ("DELTACORP",   "DELTACORP.NS",   4200),
    ("DIVISLAB",    "DIVISLAB.NS",     200),
    ("DIXON",       "DIXON.NS",        150),
    ("DLF",         "DLF.NS",         1650),
    ("DRREDDY",     "DRREDDY.NS",      250),
    ("EICHERMOT",   "EICHERMOT.NS",    175),
    ("ESCORTS",     "ESCORTS.NS",      275),
    ("EXIDEIND",    "EXIDEIND.NS",    3600),
    ("FEDERALBNK",  "FEDERALBNK.NS",  5000),
    ("GAIL",        "GAIL.NS",        6400),
    ("GLENMARK",    "GLENMARK.NS",    1150),
    ("GMRINFRA",    "GMRINFRA.NS",   22500),
    ("GNFC",        "GNFC.NS",        1200),
    ("GODREJCP",    "GODREJCP.NS",    1000),
    ("GODREJPROP",  "GODREJPROP.NS",   400),
    ("GRANULES",    "GRANULES.NS",    2700),
    ("GRASIM",      "GRASIM.NS",       475),
    ("GUJGASLTD",   "GUJGASLTD.NS",   750),
    ("HAL",         "HAL.NS",          200),
    ("HAVELLS",     "HAVELLS.NS",      500),
    ("HCLTECH",     "HCLTECH.NS",      350),
    ("HDFCAMC",     "HDFCAMC.NS",      200),
    ("HDFCBANK",    "HDFCBANK.NS",     550),
    ("HDFCLIFE",    "HDFCLIFE.NS",    1500),
    ("HEROMOTOCO",  "HEROMOTOCO.NS",   300),
    ("HINDALCO",    "HINDALCO.NS",    2150),
    ("HINDCOPPER",  "HINDCOPPER.NS",  3650),
    ("HINDPETRO",   "HINDPETRO.NS",   2700),
    ("HINDUNILVR",  "HINDUNILVR.NS",   300),
    ("HONAUT",      "HONAUT.NS",        15),
    ("IDFCFIRSTB",  "IDFCFIRSTB.NS",  7500),
    ("IEX",         "IEX.NS",         3750),
    ("IGL",         "IGL.NS",         1375),
    ("INDHOTEL",    "INDHOTEL.NS",    3000),
    ("INDIAMART",   "INDIAMART.NS",    150),
    ("INDIGO",      "INDIGO.NS",       300),
    ("INDUSINDBK",  "INDUSINDBK.NS",   500),
    ("INDUSTOWER",  "INDUSTOWER.NS",  2800),
    ("INFY",        "INFY.NS",         400),
    ("IOC",         "IOC.NS",         3500),
    ("IPCALAB",     "IPCALAB.NS",      700),
    ("IRCTC",       "IRCTC.NS",        875),
    ("ITC",         "ITC.NS",         3200),
    ("JINDALSTEL",  "JINDALSTEL.NS",  1250),
    ("JKCEMENT",    "JKCEMENT.NS",     200),
    ("JSL",         "JSL.NS",         3000),
    ("JSWENERGY",   "JSWENERGY.NS",   1500),
    ("JSWSTEEL",    "JSWSTEEL.NS",     600),
    ("JUBLFOOD",    "JUBLFOOD.NS",     625),
    ("KAJARIACER",  "KAJARIACER.NS",  1000),
    ("KOTAKBANK",   "KOTAKBANK.NS",    400),
    ("KPITTECH",    "KPITTECH.NS",     800),
    ("LALPATHLAB",  "LALPATHLAB.NS",   300),
    ("LAURUSLABS",  "LAURUSLABS.NS",  2500),
    ("LICHSGFIN",   "LICHSGFIN.NS",   1600),
    ("LT",          "LT.NS",           150),
    ("LTF",         "LTF.NS",         5000),
    ("LTIM",        "LTIM.NS",         150),
    ("LTTS",        "LTTS.NS",         200),
    ("LUPIN",       "LUPIN.NS",        500),
    ("M&M",         "M&M.NS",          175),
    ("MANAPPURAM", "MANAPPURAM.NS",   4000),
    ("MARICO",      "MARICO.NS",      1200),
    ("MARUTI",      "MARUTI.NS",       100),
    ("MCDOWELL-N",  "MCDOWELL-N.NS",  1250),
    ("MCX",         "MCX.NS",          400),
    ("METROPOLIS",  "METROPOLIS.NS",   400),
    ("MFSL",        "MFSL.NS",        1100),
    ("MGL",         "MGL.NS",          550),
    ("MOTHERSON",   "MOTHERSON.NS",   14000),
    ("MPHASIS",     "MPHASIS.NS",      300),
    ("MRF",         "MRF.NS",            5),
    ("MUTHOOTFIN",  "MUTHOOTFIN.NS",   600),
    ("NATIONALUM",  "NATIONALUM.NS",   7500),
    ("NAUKRI",      "NAUKRI.NS",       150),
    ("NAVINFLUOR",  "NAVINFLUOR.NS",   200),
    ("NESTLEIND",   "NESTLEIND.NS",     50),
    ("NMDC",        "NMDC.NS",         6750),
    ("NTPC",        "NTPC.NS",        2250),
    ("OBEROIRLTY",  "OBEROIRLTY.NS",   700),
    ("OFSS",        "OFSS.NS",          75),
    ("ONGC",        "ONGC.NS",        1925),
    ("PAGEIND",     "PAGEIND.NS",       15),
    ("PEL",         "PEL.NS",          750),
    ("PERSISTENT",  "PERSISTENT.NS",   175),
    ("PETRONET",    "PETRONET.NS",    3000),
    ("PFC",         "PFC.NS",         2700),
    ("PIDILITIND",  "PIDILITIND.NS",   250),
    ("PIIND",       "PIIND.NS",        250),
    ("PNB",         "PNB.NS",         8000),
    ("POLYCAB",     "POLYCAB.NS",      175),
    ("POWERGRID",   "POWERGRID.NS",   2700),
    ("PVRINOX",     "PVRINOX.NS",      873),
    ("RBLBANK",     "RBLBANK.NS",     5000),
    ("RECLTD",      "RECLTD.NS",      3000),
    ("RELIANCE",    "RELIANCE.NS",    1250),
    ("SAIL",        "SAIL.NS",        8500),
    ("SBICARD",     "SBICARD.NS",     1000),
    ("SBILIFE",     "SBILIFE.NS",      750),
    ("SBIN",        "SBIN.NS",        1500),
    ("SHREECEM",    "SHREECEM.NS",      25),
    ("SHRIRAMFIN",  "SHRIRAMFIN.NS",   300),
    ("SIEMENS",     "SIEMENS.NS",      275),
    ("SRF",         "SRF.NS",          375),
    ("SUNPHARMA",   "SUNPHARMA.NS",    350),
    ("SUNTV",       "SUNTV.NS",       1500),
    ("SYNGENE",     "SYNGENE.NS",     1500),
    ("TATACHEM",    "TATACHEM.NS",    1100),
    ("TATACOMM",    "TATACOMM.NS",     700),
    ("TATACONSUM",  "TATACONSUM.NS",  1100),
    ("TATAMOTORS",  "TATAMOTORS.NS",  1400),
    ("TATAPOWER",   "TATAPOWER.NS",   3375),
    ("TATASTEEL",   "TATASTEEL.NS",   5500),
    ("TCS",         "TCS.NS",          150),
    ("TECHM",       "TECHM.NS",        400),
    ("TIINDIA",     "TIINDIA.NS",      275),
    ("TITAN",       "TITAN.NS",        375),
    ("TORNTPHARM",  "TORNTPHARM.NS",   500),
    ("TORNTPOWER",  "TORNTPOWER.NS",  1500),
    ("TRENT",       "TRENT.NS",        375),
    ("TVSMOTOR",    "TVSMOTOR.NS",     350),
    ("UBL",         "UBL.NS",          700),
    ("ULTRACEMCO",  "ULTRACEMCO.NS",   200),
    ("UNIONBANK",   "UNIONBANK.NS",    8400),
    ("UPL",         "UPL.NS",         1300),
    ("VEDL",        "VEDL.NS",        2750),
    ("VOLTAS",      "VOLTAS.NS",      1000),
    ("WIPRO",       "WIPRO.NS",       1500),
    ("ZOMATO",      "ZOMATO.NS",      4500),
    ("ZYDUSLIFE",   "ZYDUSLIFE.NS",    700),
]

INDEX_LOTS = [
    ("NIFTY",      75),
    ("BANKNIFTY",  30),
    ("FINNIFTY",   40),
    ("MIDCPNIFTY", 120),
    ("NIFTYNXT50", 25),
    ("SENSEX",     20),
    ("BANKEX",     15),
]


def thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(cell, bg=NAVY, fg=WHITE, sz=9, bold=True):
    cell.font      = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = thin()

def vc(ws, row, col, value, bg=WHITE, fmt=None, bold=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9, bold=bold)
    c.fill      = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin()
    if fmt: c.number_format = fmt
    return c


def fetch_prices() -> pd.DataFrame:
    """Fetch previous day OHLC for all F&O stocks via yfinance."""
    print(f"   Fetching prices for {len(FO_STOCKS)} stocks via Yahoo Finance...")

    yahoo_symbols = [s[1] for s in FO_STOCKS]
    nse_symbols   = {s[1]: s[0] for s in FO_STOCKS}
    lot_map       = {s[1]: s[2] for s in FO_STOCKS}

    # Download 5 days to ensure we get the latest trading day
    raw = yf.download(
        yahoo_symbols,
        period="5d",
        interval="1d",
        group_by="ticker",
        auto_adjust=True,
        progress=False,
        threads=True,
    )

    rows = []
    for ysym in yahoo_symbols:
        try:
            if len(yahoo_symbols) == 1:
                df_sym = raw
            else:
                df_sym = raw[ysym] if ysym in raw.columns.get_level_values(0) else pd.DataFrame()

            if df_sym.empty or len(df_sym) < 2:
                rows.append({
                    "Symbol": nse_symbols[ysym],
                    "Lot Size": lot_map[ysym],
                    "Prev Close (₹)": 0, "Open (₹)": 0,
                    "High (₹)": 0, "Low (₹)": 0, "Close (₹)": 0,
                    "Change (₹)": 0, "Change (%)": 0,
                })
                continue

            latest   = df_sym.iloc[-1]
            prev_row = df_sym.iloc[-2]

            close      = round(float(latest["Close"]), 2)
            prev_close = round(float(prev_row["Close"]), 2)
            chg        = round(close - prev_close, 2)
            chg_pct    = round((chg / prev_close * 100) if prev_close else 0, 2)

            rows.append({
                "Symbol":        nse_symbols[ysym],
                "Lot Size":      lot_map[ysym],
                "Prev Close (₹)": prev_close,
                "Open (₹)":      round(float(latest["Open"]), 2),
                "High (₹)":      round(float(latest["High"]), 2),
                "Low (₹)":       round(float(latest["Low"]), 2),
                "Close (₹)":     close,
                "Change (₹)":    chg,
                "Change (%)":    chg_pct,
            })
        except Exception as e:
            rows.append({
                "Symbol": nse_symbols[ysym],
                "Lot Size": lot_map[ysym],
                "Prev Close (₹)": 0, "Open (₹)": 0,
                "High (₹)": 0, "Low (₹)": 0, "Close (₹)": 0,
                "Change (₹)": 0, "Change (%)": 0,
            })

    df = pd.DataFrame(rows)
    good = (df["Close (₹)"] > 0).sum()
    print(f"   → {good}/{len(rows)} stocks with price data")
    return df


def build_excel(df: pd.DataFrame, output_path: str):
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    today = datetime.now().strftime("%d-%b-%Y %I:%M %p")
    wb = Workbook(); wb.remove(wb.active)

    # ── Summary ──────────────────────────────────────────────────────────────
    ws_s = wb.create_sheet("📊 Summary")
    ws_s.sheet_view.showGridLines = False
    ws_s.merge_cells("A1:E1")
    t = ws_s.cell(row=1, column=1,
        value=f"NSE NFO Snapshot  |  Updated: {today}")
    t.font=Font(name="Arial",bold=True,size=13,color=WHITE)
    t.fill=PatternFill("solid",start_color=NAVY)
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws_s.row_dimensions[1].height=30
    for ci,h in enumerate(["Metric","Value"],1):
        hdr(ws_s.cell(row=2,column=ci),bg=GOLD,fg=NAVY)
    for ri,(k,v) in enumerate([
        ("Total F&O Stocks", len(df)),
        ("Stocks with Prices", int((df["Close (₹)"]>0).sum())),
        ("Total Indices", len(INDEX_LOTS)),
        ("Last Updated", today),
        ("Price Source", "Yahoo Finance (NSE)"),
        ("Lot Size Source", "NSE Circular — Apr 2026"),
    ],3):
        vc(ws_s,ri,1,k,LIGHT,bold=True); vc(ws_s,ri,2,v,WHITE)
    ws_s.column_dimensions["A"].width=24; ws_s.column_dimensions["B"].width=40

    ws_s.cell(row=10,column=1,value="Index F&O Lot Sizes (Jan 2026 NSE Circular)")
    hdr(ws_s.cell(row=10,column=1),bg=NAVY); ws_s.merge_cells("A10:B10")
    for ci,h in enumerate(["Index","Lot Size"],1):
        hdr(ws_s.cell(row=11,column=ci),bg=GOLD,fg=NAVY)
    for ri,(sym,lot) in enumerate(INDEX_LOTS,12):
        vc(ws_s,ri,1,sym,LIGHT,bold=True); vc(ws_s,ri,2,lot,WHITE)

    # ── F&O Stocks ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("📈 F&O Stocks")
    ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:J1")
    t=ws.cell(row=1,column=1,value=f"NSE F&O Stocks — Lot Size & OHLC  |  {today}")
    t.font=Font(name="Arial",bold=True,size=12,color=WHITE)
    t.fill=PatternFill("solid",start_color=NAVY)
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=28

    cols=["#","Symbol","Lot Size","Prev Close (₹)","Open (₹)","High (₹)",
          "Low (₹)","Close (₹)","Change (₹)","Change (%)"]
    for ci,col in enumerate(cols,1):
        ws.cell(row=2,column=ci,value=col); hdr(ws.cell(row=2,column=ci),bg=GOLD,fg=NAVY,sz=9)
    ws.row_dimensions[2].height=22
    ws.freeze_panes="A3"; ws.auto_filter.ref=f"A2:{get_column_letter(len(cols))}2"

    for ri,(_, row) in enumerate(df.sort_values("Symbol").iterrows(),3):
        bg=WHITE if ri%2==0 else LIGHT
        vc(ws,ri,1,ri-2,bg)
        vc(ws,ri,2,row["Symbol"],bg,bold=True)
        vc(ws,ri,3,int(row["Lot Size"]),bg)
        for ci,key in enumerate(["Prev Close (₹)","Open (₹)","High (₹)",
                                  "Low (₹)","Close (₹)","Change (₹)"],4):
            vc(ws,ri,ci,row.get(key,0),bg,"#,##0.00")
        chg=row.get("Change (%)",0)
        c=vc(ws,ri,10,chg,bg,"0.00")
        c.fill=PatternFill("solid",start_color=GREEN if float(chg)>=0 else ORANGE)

    widths=[5,16,10,14,14,14,14,14,12,12]
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(ci)].width=w

    # ── Indices ───────────────────────────────────────────────────────────────
    ws_i=wb.create_sheet("🏛️ Indices")
    ws_i.sheet_view.showGridLines=False
    ws_i.merge_cells("A1:C1")
    t=ws_i.cell(row=1,column=1,value=f"NSE Index F&O Lot Sizes  |  {today}")
    t.font=Font(name="Arial",bold=True,size=12,color=WHITE)
    t.fill=PatternFill("solid",start_color=NAVY)
    t.alignment=Alignment(horizontal="center",vertical="center")
    ws_i.row_dimensions[1].height=28
    for ci,h in enumerate(["#","Index","Lot Size"],1):
        hdr(ws_i.cell(row=2,column=ci),bg=GOLD,fg=NAVY)
    for ri,(sym,lot) in enumerate(INDEX_LOTS,3):
        bg=WHITE if ri%2==0 else LIGHT
        vc(ws_i,ri,1,ri-2,bg); vc(ws_i,ri,2,sym,bg,bold=True); vc(ws_i,ri,3,lot,bg)
    ws_i.column_dimensions["A"].width=5
    ws_i.column_dimensions["B"].width=20
    ws_i.column_dimensions["C"].width=12

    # ── Change Log ─────────────────────────────────────────────────────────────
    ws_log=wb.create_sheet("📋 Change Log")
    for ci,h in enumerate(["Run Date","Stocks","With Prices","Indices","Source"],1):
        hdr(ws_log.cell(row=1,column=ci)); ws_log.cell(row=1,column=ci).value=h
    ws_log.append([today,len(df),int((df["Close (₹)"]>0).sum()),
                   len(INDEX_LOTS),"Yahoo Finance + NSE Circular"])
    for w,col in zip([22,10,12,10,35],["A","B","C","D","E"]):
        ws_log.column_dimensions[col].width=w

    wb.save(output_path)
    print(f"✅ Excel saved → {output_path}")


def main():
    print("🚀 NSE NFO Lot Size & Price Tracker")
    print(f"   Stocks in list: {len(FO_STOCKS)}")
    print(f"   Indices: {len(INDEX_LOTS)}")

    print("\n📈 Fetching prices from Yahoo Finance...")
    df = fetch_prices()

    out_path = os.path.abspath(OUT)
    print(f"\n📊 Building Excel...")
    build_excel(df, out_path)

    snap = {
        "updated_at": datetime.now().isoformat(),
        "total_stocks": len(df),
        "stocks_with_prices": int((df["Close (₹)"] > 0).sum()),
        "source": "Yahoo Finance (yfinance)",
    }
    with open(out_path.replace(".xlsx", ".json"), "w") as f:
        json.dump(snap, f, indent=2, default=str)

    print("🎉 Done!")


if __name__ == "__main__":
    main()
